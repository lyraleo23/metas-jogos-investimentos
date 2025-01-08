import os
import datetime
import openpyxl


def main():
    """
    Função principal que gerencia a planilha de metas de jogos.
    A função verifica se o arquivo Excel "metas_jogos.xlsx" existe. Se não existir, cria um novo workbook.
    Em seguida, verifica a quantidade de linhas na planilha e adiciona os dados de metas diárias e dívidas.
    Se a data atual já estiver registrada, não adiciona novos dados.
    Variáveis:
    - meta_diaria (int): Meta diária de troféus.
    - meta_trofeus (int): Total de troféus acumulados.
    - divida (int): Dívida acumulada.
    - data_atual (str): Data atual no formato "dd/mm/yyyy".
    Exceções:
    - FileNotFoundError: Se o arquivo "metas_jogos.xlsx" não for encontrado, cria um novo workbook.
    Ações:
    - Adiciona dados na planilha se a data atual for diferente da última registrada.
    - Salva o arquivo Excel com os dados atualizados.
    Prints:
    - Informa a quantidade de linhas na planilha.
    - Informa se a data atual já foi registrada.
    - Informa o sucesso ao salvar a planilha com as metas e dívidas atualizadas.
    """

    os.system("cls" if os.name == "nt" else "clear")
    print("Bem-vindo ao Gerenciador de Metas de Jogos!")

    # Verifica se o workbook existe
    try:
        workbook = openpyxl.load_workbook("metas_jogos.xlsx")
    except FileNotFoundError:
        criar_workbook()
        workbook = openpyxl.load_workbook("metas_jogos.xlsx")

    # Variáveis
    meta_diaria = 5
    meta_trofeus = 0
    divida = 0
    data_atual = datetime.datetime.now().strftime("%d/%m/%Y")
    
    # Verifica quantas linhas tem na planilha
    sheet = workbook.active
    num_linhas = sheet.max_row

    if num_linhas == 1:
        # Adiciona os dados na planilha
        sheet["A2"] = data_atual
        sheet["B2"] = meta_trofeus + meta_diaria
        sheet["C2"] = divida
    elif num_linhas > 1:
        # Verifica se a data atual é diferente da última data
        if sheet[f"A{num_linhas}"].value != data_atual:
            ultima_data = datetime.datetime.strptime(sheet[f"A{num_linhas}"].value, "%d/%m/%Y")
            data_atual_dt = datetime.datetime.strptime(data_atual, "%d/%m/%Y")
            diferenca_dias = (data_atual_dt - ultima_data).days
            print(f"A diferença em dias é: {diferenca_dias}")

            meta_trofeus = sheet[f"B{num_linhas}"].value + (meta_diaria * diferenca_dias)
            divida = sheet[f"C{num_linhas}"].value
            num_linhas += 1

            # Adiciona os dados na planilha
            sheet[f"A{num_linhas}"] = data_atual
            sheet[f"B{num_linhas}"] = meta_trofeus
            sheet[f"C{num_linhas}"] = divida
        else:
            print("A data atual já foi registrada na planilha.")

    print(f'Meta atual: {sheet[f"B{num_linhas}"].value} Troféus. Dívida atual1: {sheet[f"C{num_linhas}"].value}.')

    num_linhas = atualizar_trofeus(sheet, num_linhas)

    # Salva o arquivo Excel
    workbook.save("metas_jogos.xlsx")
    print(f'Planilha salva com sucesso. Meta: {sheet[f"B{num_linhas}"].value} Troféus. Dívida: {sheet[f"C{num_linhas}"].value}.')
    print("Obrigado por usar o Gerenciador de Metas de Jogos!")

def criar_workbook():
    # Cria um novo workbook e seleciona a planilha ativa
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Adiciona os dados na planilha
    sheet["A1"] = "Data Atual"
    sheet["B1"] = "Meta Troféus"
    sheet["C1"] = "Dívida"

    # Salva o arquivo Excel
    workbook.save("metas_jogos.xlsx")

def atualizar_trofeus(sheet, num_linhas):
    """
    Função para atualizar a quantidade de troféus ganhos pelo usuário.
    Para cada troféu ganho, diminui a quantidade de meta_trofeus e aumenta a dívida em 10.
    """
    trofeus_ganhos = int(input("Quantos troféus foram ganhos hoje? "))
    meta_trofeus = sheet[f"B{num_linhas}"].value
    divida = sheet[f"C{num_linhas}"].value

    for _ in range(trofeus_ganhos):
        if meta_trofeus > 0:
            meta_trofeus -= 1
            divida += 10

    num_linhas += 1
    sheet[f"A{num_linhas}"] = datetime.datetime.now().strftime("%d/%m/%Y")
    sheet[f"B{num_linhas}"] = meta_trofeus
    sheet[f"C{num_linhas}"] = divida
    print(f'Atualização: Meta: {meta_trofeus} Troféus. Dívida: {divida}.')

    return num_linhas


main()
